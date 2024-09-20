import time
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.workbook import Workbook

data = pd.read_excel('读取文件名')

wb = Workbook()
ws = wb.active

font_style = Font(size=12, bold=True)

fill_style = PatternFill(start_color="92D050",
                         end_color="92D050",
                         fill_type="solid",
                         )

time.sleep(0.5)
alignment_style = Alignment(horizontal='center', vertical='center')

# 此处是输出表头
xlsx_headers = ['测试表头']
for col_num, column_title in enumerate(xlsx_headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = column_title
    cell.font = font_style
    cell.fill = fill_style
    cell.alignment = alignment_style
    ws.column_dimensions[chr(64+col_num)].width = 30

ws.row_dimensions[1].height = 25



for index, row in data.iterrows():
  ## todo 业务逻辑

    ws.append(row.to_list())


for i in range(1, ws.max_row + 1):
    for j in range(1, ws.max_column + 1):
        ws.cell(i, j).alignment = alignment_style

wb.save('保存文件名')
time.sleep(10)







